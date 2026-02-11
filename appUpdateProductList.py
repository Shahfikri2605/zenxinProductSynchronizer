import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter 
from copy import copy
import io
from collections import defaultdict

st.set_page_config(page_title="Product List Sync (Graded Quantity)", layout="wide")
st.title("📦 Product List Synchronizer")

LOCATION_MAPPING = {
    # --- Exact / High Confidence Matches ---
    "Aeon Bukit Indah":"Aeon Bukit Indah-JHR",
    "Aeon Tebrau City": "Aeon Tebrau-JHR",
    "Aeon Bandar Dato Onn": "Aeon Dato Onn-JHR",
    "Aeon Bandar Utama": "Aeon Bandar Utama-KUL",
    "Aeon Bukit Tinggi": "Aeon Bukit Tinggi-KUL",
    "Aeon Cheras Selatan Balakong": "2 Aeon Cheras Selatan-KUL",
    "Aeon KulaiJaya": "Aeon Kulai Jaya-JHR",
    "Aeon Maxvalu Danga Bay": "Aeon Maxvalu Danga Bay",
    "Aeon Permas Jaya": "Aeon Permas-JHR",
    "Aeon Putrajaya": "Aeon Putrajaya-KUL",
    "Aeon Seremban 2": "Aeon Seremban 2-KUL",
    "Aeon Taman Maluri": "AM Aeon Taman Maluri-KUL",
    "Aeon Taman Universiti": "Aeon Taman U-JHR",
    "Aeon Mid Valley": "Aeon Midvalley-KUL",
    "Aeon IOI Putrajaya": "Aeon IOI Puchong-KUL",
    "Aeon Putrajaya": "Aeon Putrajaya-KUL ",
    "Aeon Metro Prima": "Aeon Metro Prima-KUL",
    "Aeon Taman Equine": "Aeon Taman Equine-KUL",
    "Aeon Nilai": "Aeon Nilai-KUL",
    "Aeon Setia Alam": "Aeon Setia Alam-KUL",
    "Aeon Shah Alam Store": "Aeon Shah Alam-KUL",
    "Aeon Wangsa Maju" : "Aeon Wangsa Maju-KUL",
    "Aeon Maxvalue Desa Waterpark City" : "Aeon Maxvalu DPC-KUL",


    "Urban Fresh MarketPlace": "Urban Fresh-KUL",

    
    "Bens - Batai": "VG Ben's Batai (BBT)-KUL",
    "Bens - IPC": "VG Ben's Ipc (BIP)-KUL",
    "Bens - Linc Bandar Tun Razak": "VG Ben's Linc (BLI)-KUL",
    "Bens - Mall Of Medini-Legoland": "VG Ben's Mall (BMM)-JHR",
    "Bens Bangsar Shopping Centre (BSC)-KUL": "VG Ben's (BSC)-KUL",
    
    "GCH - Giant Leisure Mall-JHR": "GCH Leisure Mall-JHR",
    "GCH - Plaza Pelangi": "GCH Plaza Pelangi-JHR",
    
    "Isetan KLCC (Consignment)": "Isetan Of Japan-KUL",
    
    "Jaya Grocer - Eco Galleria-JHR": "JG Eco Galleria-JHR",
    "Jaya Grocer - Intermark": "JG Intermark-KUL",
    "Jaya Grocer - Iskandar Johor": "JG Iskandar-JHR",
    "Jaya Grocer - Mutiara Tropicana": "JG Mutiara Tropicana-KUL",
    "Jaya Grocer - Rio Ioi Puchong": "2 JG Rio Puchong-KUL",
    "Jaya Grocer - Seremban Prima": "JG Prima Seremban-KUL",
    "Jaya Grocer - Sunway Pyramid": "JG Sunway Pyramid-KUL",
    "Jaya Grocer - Tropica Bukit Jalil": "JG Tropika Bukit Jalil-KUL",
    "Jaya Grocer - KLGCC Mall" : "JG KLGCC Mall-KUL",
    "Jaya Grocer - Eco Majestic" : "JG Eco Majestic-KUL",
    "Jaya Grocer - Subang Jaya Empire" : "JG Subang Empire-KUL",
    "Jaya Grocer - USJ21" : "JG USJ-KUL",
    "Jaya Grocer - Verve Mont Kiara" : "JG Verve Mont Kiara-KUL",
    "Jaya Grocer - One Utama": "JG 1 Utama-KUL",
    "Jaya Grocer - Cyberjaya": "JG Cyberjaya-KUL",
    "Jaya Grocer - Elmina Lakeside Mall-KUL" : "JG Elmina-KUL",
    "Jaya Grocer - Bukit Jelutong": "JG Jelutong-KUL",

    
    "TFM Pavilion Bukit Bintang-KUL": "TFM Pavilion Kuala Lumpur-KUL",
    "TFM Pavillion Embassy-KUL": "TFM Pavilion Embassy-KUL",
    "TFM WCity OUG Sales Gallery-KUL":"TFM WCity OUG Sales Gallery-KUL",

    
    "Village Grocer - Avenue K Ampang": "VG Avenue K (VAK)-KUL",
    "Village Grocer - Cheras Leisure Mall (LGC)-KUL": "VG Leisure Mall (LGC)-KUL",
    "Village Grocer - Eco Cheras": "VG Eko Cheras (VEC)-KUL",
    "Village Grocer - Kota Damansara Giza": "VG Giza (VGG)-KUL",
    "Village Grocer - Bangsar Village": "VG Bangsar Village (VGB)-KUL",
    "Village Grocer - Desa Park City Plaza Arkadia": "VG Desa Park City (VDP)-KUL",
    "Village Grocer - Damansara Jaya Atria": "VG Atria (VDJ)-KUL",
    "Bens - Publika" : "VG Ben's Publika (BPS)-KUL",
    "Village Grocer - Citta Mall Ara Damansara": "VG Citta Mall (VAD)-KUL",
    "Village Grocer - Laman Seri Harmoni 33 (VLH)": "VG Laman Seri Harmoni 33 (VLH)-KUL",
    "Bens - Batai": "VG Ben's Batai (BBT)-KUL",
    "Bens Bangsar Shopping Centre (BSC)-KUL": "VG Ben's (BSC)-KUL",
    "Bens - IPC": "VG Ben's Ipc (BIP)-KUL",
    "Bens - Linc Bandar Tun Razak": "VG Ben's Linc (BLI)-KUL",
    "Village Grocer - Mont Kiara": "VG Mont Kiara (VGO)-KUL",
    "Village Grocer - Subang Parade": "VG Subang Parade (VSP)-KUL",
    "Village Grocer - Tamarind Square Cyberjaya": "Village Grocer - Tamarind Square Cyberjaya",
    "Village Grocer - Myra Park Marketplace-KUL": "VG Myra Park Marketplace-KUL ",
    "Village Grocer - Puchong": "VG Puchong-KUL",
    
}


def clean_text_strict(text):
    if not isinstance(text, str): return set()
    text = text.lower()
    # Keep Country codes, remove generic units
    remove_list = ["-jhr", "-kul", "-png", "(m)", "packet", "pck", "pcs", "kg", "g", "ea", "org"]
    for word in remove_list:
        text = text.replace(f" {word} ", " ") 
        text = text.replace(f" {word}", " ")
        text = text.replace(f"{word} ", " ")
    for char in "()-,.":
        text = text.replace(char, " ")
    return set(text.split())

def find_best_column_match_strict(sheet, header_row_idx, target_location):
    """
    100% Strict Location Matching.
    1. Check Mapping.
    2. Look for EXACT string in header (ignoring spaces).
    """
    # 1. Translate using Map
    if target_location in LOCATION_MAPPING:
        search_name = LOCATION_MAPPING[target_location].strip().lower()
    else:
        search_name = target_location.strip().lower()

    # 2. Iterate headers and find EXACT match
    for cell in sheet[header_row_idx]:
        if not cell.value: continue
        
        header_val = str(cell.value).strip().lower()
        
        if header_val == search_name:
            return cell.column
            
    return None # No match if not found exactly

def find_best_product_match_100_percent(target_name, available_names):
    target_words = clean_text_strict(target_name)
    if not target_words: return None
    best_match = None
    best_fit_score = 0 

    for name in available_names:
        candidate_words = clean_text_strict(name)
        if target_words.issubset(candidate_words):
            if len(candidate_words) == 0: continue
            score = len(target_words) / len(candidate_words)
            if score > best_fit_score:
                best_fit_score = score
                best_match = name
    return best_match

def get_default_qty_by_grade(sheet, col_idx, header_row_idx):
    try:
        grade_row = header_row_idx - 1
        if grade_row < 1: return 10 
        cell_val = sheet.cell(row=grade_row, column=col_idx).value
        grade = str(cell_val).strip().upper()
        if 'A' in grade: return 10
        if 'B' in grade: return 8
        if 'C' in grade: return 4
        return 10 
    except:
        return 10

# --- MAIN APP ---

c1, c2, c3 = st.columns(3)
master_file = c1.file_uploader("1. Master List (Pivot)", type=['xlsx'])
daily_file = c2.file_uploader("2. Daily Sheet (Target)", type=['xlsx'])
report_file = c3.file_uploader("3. Request/Reduce Report", type=['xlsx', 'csv'])

if master_file and daily_file:
    # Load Master
    df_master_raw = pd.read_excel(master_file, header=None)
    master_data = []
    for _, row in df_master_raw.iterrows():
        code, name = str(row[0]).strip(), str(row[1]).strip()
        if code.lower() not in ['nan', 'none', '']:
            master_data.append({'Item Code': code, 'Master_Name': name})
    df_master = pd.DataFrame(master_data).drop_duplicates('Item Code')

    # Load Daily
    wb = openpyxl.load_workbook(daily_file)
    selected_sheet = st.selectbox("Select Sheet to Process:", wb.sheetnames)

    if st.button("Run Sync & Fix Formulas"):
        sheet = wb[selected_sheet]
        
        # Detect Header
        data_start_row = 7
        header_row = 6
        for row in sheet.iter_rows(min_row=1, max_row=20):
            val = str(row[1].value).strip() if row[1].value else ""
            if val.startswith('BP') or val.startswith('100'):
                data_start_row = row[0].row
                header_row = data_start_row - 1
                break
        
        # Find last column for SUM formula
        max_col_letter = get_column_letter(sheet.max_column)

        # 1. READ ORIGINAL DATA & STYLES
        original_rows_data = {}
        default_style_row = [] 

        for row in sheet.iter_rows(min_row=data_start_row):
            item_code = str(row[1].value).strip()
            row_styles = []
            for cell in row:
                row_styles.append({
                    'value': cell.value,
                    'fill': copy(cell.fill),
                    'font': copy(cell.font),
                    'border': copy(cell.border),
                    'alignment': copy(cell.alignment),
                    'number_format': cell.number_format
                })
            if not default_style_row: default_style_row = row_styles
            if item_code not in original_rows_data:
                original_rows_data[item_code] = row_styles

        # 2. CREATE NEW ROW ORDER
        final_list_of_row_styles = []
        for m_code in df_master['Item Code']:
            if m_code in original_rows_data:
                final_list_of_row_styles.append(original_rows_data[m_code])
            else:
                new_row_style = [copy(s) for s in default_style_row]
                for i in range(len(new_row_style)): new_row_style[i]['value'] = None
                new_row_style[0]['value'] = m_code 
                new_row_style[1]['value'] = m_code
                m_name = df_master[df_master['Item Code'] == m_code]['Master_Name'].values[0]
                new_row_style[3]['value'] = m_name 
                final_list_of_row_styles.append(new_row_style)

        for d_code, d_styles in original_rows_data.items():
            if d_code not in df_master['Item Code'].values:
                final_list_of_row_styles.append(d_styles)

        # 3. WRITE BACK (WITH FORMULA FIX)
        for row in sheet.iter_rows(min_row=data_start_row):
            for cell in row: cell.value = None

        product_row_map = defaultdict(list) 

        for r_idx, styled_row in enumerate(final_list_of_row_styles, start=data_start_row):
            p_name = str(styled_row[3]['value']).strip()
            product_row_map[p_name].append(r_idx)
            
            for c_idx, s in enumerate(styled_row, start=1):
                cell = sheet.cell(row=r_idx, column=c_idx)
                
                # --- FORMULA INJECTION ---
                if c_idx == 6: # MY Used
                    cell.value = f"=SUM(J{r_idx}:{max_col_letter}{r_idx})"
                elif c_idx == 7: # Diff
                    cell.value = f"=E{r_idx}-F{r_idx}"
                elif c_idx == 9: # Prepared Diff
                    cell.value = f"=H{r_idx}-G{r_idx}"
                else:
                    val = s['value']
                    cell.value = None if pd.isna(val) or str(val).lower() in ['nan', 'none'] else val
                
                cell.fill = s['fill']
                cell.font = s['font']
                cell.border = s['border']
                cell.alignment = s['alignment']
                cell.number_format = s['number_format']

        # 4. REPORT UPDATE
        if report_file:
            try:
                if report_file.name.endswith('.csv'):
                    df_report = pd.read_csv(report_file)
                else:
                    df_report = pd.read_excel(report_file)

                df_req = df_report.iloc[:, 0:5].copy()
                df_red = df_report.iloc[:, 6:11].copy()
                
                yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

                available_products = list(product_row_map.keys())

                # Request Loop
                df_req = df_req.dropna(subset=['Veggie Request'])
                for _, row in df_req.iterrows():
                    loc_name = str(row['Location'])
                    raw_prod_name = str(row['Veggie Request']).strip()
                    qty = row['Qty']
                    
                    official_prod = find_best_product_match_100_percent(raw_prod_name, available_products)
                    col_idx = find_best_column_match_strict(sheet, header_row, loc_name)

                    if official_prod and col_idx:
                        if pd.isna(qty) or str(qty).strip() == "":
                            qty_to_write = get_default_qty_by_grade(sheet, col_idx, header_row)
                        else:
                            qty_to_write = qty

                        for row_idx in product_row_map[official_prod]:
                            cell = sheet.cell(row=row_idx, column=col_idx)
                            cell.value = qty_to_write
                            cell.fill = yellow_fill
                    else:
                        if not official_prod: st.warning(f"Request Mismatch: '{raw_prod_name}'")
                        elif not col_idx: st.warning(f"Location Mismatch: '{loc_name}'")

                # Reduce Loop
                df_red = df_red.dropna(subset=['Veggie Reduce'])
                for _, row in df_red.iterrows():
                    loc_name = str(row['Location.1'])
                    raw_prod_name = str(row['Veggie Reduce']).strip()
                    
                    official_prod = find_best_product_match_100_percent(raw_prod_name, available_products)
                    col_idx = find_best_column_match_strict(sheet, header_row, loc_name)

                    if official_prod and col_idx:
                        for row_idx in product_row_map[official_prod]:
                            cell = sheet.cell(row=row_idx, column=col_idx)
                            cell.fill = red_fill
                    else:
                        if not official_prod: st.warning(f"Reduce Mismatch: '{raw_prod_name}'")
                        elif not col_idx: st.warning(f"Location Mismatch: '{loc_name}'")
                
                st.success("Process Complete! Formulas restored and Graded Quantities applied.")

            except Exception as e:
                st.error(f"Report Error: {e}")

        output = io.BytesIO()
        wb.save(output)
        st.download_button("Download Updated File", output.getvalue(), f"Updated_File{selected_sheet}.xlsx")
